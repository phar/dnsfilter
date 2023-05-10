import openpyxl
import socket
import sys

def resolve_hosts_from_excel(file_path, output_file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the column index for the "Domain" column
    domain_column_index = None
    for column in sheet.iter_cols():
        if column[0].value == "Domain":
            domain_column_index = column[0].column
            break

    if domain_column_index is None:
        print("Domain column not found in the Excel file.")
        return

    # Extract hostnames and URLs from the "Domain" column
    hostnames = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        domain_value = row[domain_column_index - 1]
        if domain_value:
            if domain_value.startswith("http://") or domain_value.startswith("https://"):
                # Extract hostname from URL
                hostname = domain_value.split("//")[-1].split("/")[0]
                hostnames.append(hostname)
            else:
                hostnames.append(domain_value)

    # Resolve hostnames
    resolved_hosts = []
    for hostname in hostnames:
        try:
            ip_address = socket.gethostbyname(hostname)
            resolved_hosts.append(f"{hostname} - {ip_address}")
        except socket.gaierror:
#            resolved_hosts.append(f"{hostname} - Unable to resolve")
			print(f"{hostname} - Unable to resolve")

    # Write resolved hosts to a file
    with open(output_file_path, "w") as file:
        for resolved_host in resolved_hosts:
            file.write(resolved_host + "\n")

    print(f"Resolved hosts saved to {output_file_path}.")

# Usage example
input_file_path = sys.argv[1]
output_file_path = "hosts.txt"

resolve_hosts_from_excel(input_file_path, output_file_path)

